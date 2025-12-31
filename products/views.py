from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from orders.models import Order, OrderItem, OrderAddress
from django.core.paginator import Paginator
from django.db.models import Sum
from datetime import datetime, timedelta
from django.utils import timezone
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from django.http import HttpResponse
from io import BytesIO
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from accounts.utils import profile
from products.models import Referral, WishlistItem, Review
from adminpanel.models import Product, Variant
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from accounts.utils import info_notify, error_notify, success_notify
from cart.models import Cart, CartItem, Wallet
from django.urls import reverse
from django.db import models

import logging

logger = logging.getLogger(__name__)
# Create your views here.

@method_decorator(login_required(login_url='signin'), name='dispatch')
@method_decorator(never_cache, name='dispatch')
class AdminSalesView(View):

    def get(self, request):
        if request.user.is_authenticated:
            if not request.user.is_superuser:
                
                return redirect("cores-home", user_id=request.user.id)
        filter_type = request.GET.get('filter')
        custom_start = request.GET.get('start_date')
        custom_end = request.GET.get('end_date')

        

        orders = Order.objects.exclude(status__in=['Cancelled', 'Payment Failed']).order_by('-created_at')

        if filter_type == 'daily':
            start_date = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timedelta(days=1)
        elif filter_type == 'weekly':
            start_date = timezone.now() - timedelta(days=timezone.now().weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timedelta(days=7)
            
        elif filter_type == 'monthly':
            today = timezone.now()
            start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = (start_date + timedelta(days=32)).replace(day=1)
            
        elif filter_type == 'yearly':
            today = timezone.now()
            start_date = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date.replace(month=12, day=31, hour=23, minute=59, second=59)
        else:
            start_date = timezone.now() - timedelta(days=30)
            end_date = timezone.now()

        if custom_start and custom_end:
            start_date = datetime.strptime(custom_start, '%Y-%m-%d')
            end_date = datetime.strptime(custom_end, '%Y-%m-%d') + timedelta(days=1)
            start_date = timezone.make_aware(start_date)
            end_date = timezone.make_aware(end_date)

        orders = orders.filter(created_at__gte=start_date, created_at__lt=end_date)
        

        total_orders = orders.count()
        total_items = sum(item.quantity for order in orders for item in order.items.filter(is_cancel=False, is_return=False))
        total_revenue = sum(order.over_all_amount for order in orders )
        avg_order_value = total_revenue / total_orders if total_orders else 0


        period_str = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"

        report_data = {
            'store_name': 'NoteVia',
            'period': period_str,
            'generated_at': timezone.now().strftime('%d/%m/%Y, %H:%M:%S'),
            'orders': orders,
            'total_orders': total_orders,
            'total_items': total_items,
            'total_revenue': round(total_revenue, 2),
            # 'total_commission': round(total_commission, 2),
            # 'vendor_payout': round(total_revenue - total_commission, 2),
            'avg_order': round(avg_order_value, 2),
        }
        
        if 'download_pdf' in request.GET:
            
            return self.generate_pdf_reportlab(report_data)
        if 'download_excel' in request.GET:
            return self.generate_excel(report_data)

        delivered_orders = Order.objects.filter(status='Delivered')
        shippped_orders = Order.objects.filter(status='Shipped')
        processing_orders = Order.objects.filter(status='Processing')

        
        
        delivered_count = delivered_orders.count()
        shipped_count = shippped_orders.count()
        processing_count  = processing_orders.count()
        orders_count = orders.count()
        
        total_quantity_sold = OrderItem.objects.filter( order__status='Delivered',is_cancel=False
                                                       ).aggregate(total=Sum('quantity'))['total'] or 0
        

        total_sales = sum(order.over_all_amount_all for order in delivered_orders)
        

        page_number = request.GET.get('page', 1)
        paginator = Paginator(orders, 6)
        page_obj = paginator.get_page(page_number)

        

        context = {
            **report_data,
            'filter_type': filter_type,
            "user_id": request.user.id,
            "orders": page_obj,
            "page_obj": page_obj,
            "delivered_count": delivered_count,
            "shipped_count": shipped_count,
            "processing_count": processing_count,
            "orders_count": orders_count,
            "total_quantity_sold": total_quantity_sold,
            "total_sales": total_sales,
            "custom_start" : custom_start,
            "custom_end": custom_end,
            "start_date": start_date,
            "end_date": end_date,


        }
        return render(request, 'products/admin_sales.html', context)
    
    def generate_pdf_reportlab(self, data):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        elements.append(Paragraph(f"<font size=18><b>ADMIN SALES REPORT - {data['store_name']}</b></font>", styles["Title"]))
        elements.append(Paragraph(f"Period: {data['period']}", styles["Normal"]))
        elements.append(Paragraph(f"Generated at: {data['generated_at']}", styles["Normal"]))
        elements.append(Spacer(1, 20))

        for index, order in enumerate(data['orders'], start=1):
            # Order Header
            elements.append(Paragraph(
                f"<b>SL.No:</b> {index} | "
                f"<b>Order ID:</b> {order.order_id} | "
                f"<b>Date:</b> {order.created_at.strftime('%d/%m/%Y %H:%M')} | "
                f"<b>Customer:</b> {order.user.full_name or order.user.email} | "
                f"<b>Payment:</b> {order.get_payment_method_display()} | "
                f"<b>Status:</b> {order.get_status_display()} |"
                f"<b>Order Total:</b> {order.over_all_amount_all}",
                styles["Normal"]
            ))
            elements.append(Spacer(1, 8))

            # Items Table Header
            item_table_data = [
                ['#', 'Product', 'Variant', 'Qty', 'Price', 'Discount Price \n per Qty', 'Line Total', 'Return \n Status', 'Cancel \n Status']
            ]
            

            total_items_amount = 0
            for idx, item in enumerate(order.items.all(), start=1):
                if item.product and item.variant:
                    product_name = item.product.name
                    variant_details = str(item.variant)  # assuming __str__ on Variant is good
                elif item.product:
                    product_name = item.product.name
                    variant_details = "-"
                else:
                    product_name = "Deleted Product"
                    variant_details = "-"

                line_total = item.quantity * item.price
                total_items_amount += line_total

                discount_display = f"Rs. {item.discount_price}" if item.discount_price else "-"
                r_status = "Return" if item.is_return else "-"
                c_status = "Cancel" if item.is_cancel else "-"

                item_table_data.append([
                    str(idx),
                    product_name,
                    variant_details,
                    str(item.quantity),
                    f"Rs. {item.real_price}",
                    discount_display,
                    f"Rs. {line_total:.2f}",
                    r_status,
                    c_status
                ])

            # Create Items Table
            item_table = Table(item_table_data, colWidths=[25, 130, 80, 40, 65, 70, 70, 35, 35])
            item_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16213e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(item_table)

            # Order Totals Section
            elements.append(Spacer(1, 12))

            # Right-aligned totals
            totals_data = [
                ["", "", "", "", "", "Subtotal:", f"Rs. {total_items_amount:.2f}"],
            ]

            if order.coupon_code and order.coupon_amount:
                totals_data.append(["", "", "", "", "", f"Coupon ({order.coupon_code}):", f"- Rs. {order.coupon_amount:.2f}"])

            wallet, created = Wallet.objects.get_or_create(user=order.user)
            if wallet.return_order_amount(order):
                totals_data.append(["", "", "", "", "", "Return amount :", f" Rs. {wallet.return_order_amount(order):.2f}"])

            grand_total = order.over_all_amount or 0
            totals_data.append(["", "", "", "", "", "Grand Total:", f"Rs. {grand_total:.2f}"])

            totals_table = Table(totals_data, colWidths=[30, 150, 100, 40, 70, 100, 80])
            totals_table.setStyle(TableStyle([
                ('ALIGN', (-2, -1), (-1, -1), 'RIGHT'),  # Align last two columns right
                ('ALIGN', (-2, 0), (-2, -1), 'RIGHT'),   # Label right-aligned
                ('GRID', (0, 0), (-1, -1), 0, colors.white),  # No grid
                ('FONTSIZE', (0, 0), (-1, -1), 11),
            ]))
            elements.append(totals_table)

            elements.append(Spacer(1, 20))  # Space between orders

        # Final Summary at the bottom
        elements.append(Paragraph("<font size=14><b>REPORT SUMMARY</b></font>", styles["Heading2"]))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"Total Orders: {data['total_orders']}", styles["Normal"]))
        elements.append(Paragraph(f"Total Items Sold: {data['total_items']}", styles["Normal"]))
        elements.append(Paragraph(f"Average Order Value: Rs. {data['avg_order']}", styles["Normal"]))
        elements.append(Paragraph(f"<b>TOTAL REVENUE: Rs. {data['total_revenue']}</b>", styles["Normal"]))

        # Build PDF
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="detailed_sales_report.pdf"'
        return response
    
    def generate_excel(self, data):
        wb = Workbook()
        ws = wb.active
        ws.title = "Detailed Sales Report"

        # Minimal Styles
        header_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right")
        left_align = Alignment(horizontal="left")

        light_gray_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        summary_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        row_idx = 1

        # Main Title
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
        title_cell = ws.cell(row=row_idx, column=1)
        title_cell.value = f"ADMIN SALES REPORT - {data['store_name']}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = center_align
        row_idx += 2

        # Period and Generated At
        ws.cell(row=row_idx, column=1).value = f"Period: {data['period']}"
        row_idx += 1
        ws.cell(row=row_idx, column=1).value = f"Generated at: {data['generated_at']}"
        row_idx += 3  # Extra space before first order

        # Loop through each order
        for index, order in enumerate(data['orders'], start=1):
            # Order Header (Light gray background)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
            header_cell = ws.cell(row=row_idx, column=1)
            header_cell.value = (
                f"SL.No: {index} | Order ID: {order.order_id} | "
                f"Date: {order.created_at.strftime('%d/%m/%Y %H:%M')} | "
                f"Customer: {order.user.full_name or order.user.email} | "
                f"Payment: {order.get_payment_method_display()} | "
                f"Status: {order.get_status_display()} | "
                f"Order Total: ₹{order.over_all_amount_all or 0:.2f}"
            )
            header_cell.font = Font(bold=True, size=12)
            header_cell.fill = light_gray_fill
            header_cell.alignment = left_align
            row_idx += 2

            # Item Table Headers (Dark blue)
            item_headers = ['#', 'Product', 'Variant', 'Qty', 'Real Price (per unit)', 
                            'Discount Price (per unit)', 'Line Total', 'Return Status', 'Cancel Status']
            for col_num, header in enumerate(item_headers, 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            row_idx += 1

            # Item Rows (Plain, no borders or fills)
            total_items_amount = 0
            items = order.items.all()
            for idx, item in enumerate(items, start=1):
                if item.product and item.variant:
                    product_name = item.product.name
                    variant_details = str(item.variant)
                elif item.product:
                    product_name = item.product.name
                    variant_details = "-"
                else:
                    product_name = "Deleted Product"
                    variant_details = "-"

                line_total = item.quantity * item.price
                total_items_amount += line_total

                discount_display = f"₹{item.discount_price:.2f}" if item.discount_price else "-"
                return_status = "Returned" if item.is_return else "-"
                cancel_status = "Cancelled" if item.is_cancel else "-"

                row_data = [
                    idx,
                    product_name,
                    variant_details,
                    item.quantity,
                    f"₹{item.real_price:.2f}",
                    discount_display,
                    f"₹{line_total:.2f}",
                    return_status,
                    cancel_status
                ]

                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.value = value
                    # Alignment: center for index/qty/status, right for prices, left for names
                    if col_num in [1, 4, 8, 9]:  # #, Qty, Return, Cancel
                        cell.alignment = center_align
                    elif col_num in [5, 6, 7]:  # Prices
                        cell.alignment = right_align
                    else:
                        cell.alignment = left_align
                row_idx += 1

            # Order Totals (Plain, right-aligned)
            row_idx += 1  # Blank line

            ws.cell(row=row_idx, column=6).value = "Subtotal:"
            ws.cell(row=row_idx, column=6).font = bold_font
            ws.cell(row=row_idx, column=6).alignment = right_align
            ws.cell(row=row_idx, column=7).value = f"₹{total_items_amount:.2f}"
            ws.cell(row=row_idx, column=7).font = bold_font
            row_idx += 1

            if order.coupon_code and order.coupon_amount:
                ws.cell(row=row_idx, column=6).value = f"Coupon ({order.coupon_code}):"
                ws.cell(row=row_idx, column=6).font = bold_font
                ws.cell(row=row_idx, column=6).alignment = right_align
                ws.cell(row=row_idx, column=7).value = f"- ₹{order.coupon_amount:.2f}"
                ws.cell(row=row_idx, column=7).font = bold_font
                row_idx += 1

            # Return amount
            try:
                wallet, _ = Wallet.objects.get_or_create(user=order.user)
                return_amount = wallet.return_order_amount(order)
                if return_amount and return_amount > 0:
                    ws.cell(row=row_idx, column=6).value = "Return Amount:"
                    ws.cell(row=row_idx, column=6).font = bold_font
                    ws.cell(row=row_idx, column=6).alignment = right_align
                    ws.cell(row=row_idx, column=7).value = f"₹{return_amount:.2f}"
                    ws.cell(row=row_idx, column=7).font = bold_font
                    row_idx += 1
            except:
                pass

            # Grand Total
            grand_total = order.over_all_amount or 0
            ws.cell(row=row_idx, column=6).value = "Grand Total:"
            ws.cell(row=row_idx, column=6).font = Font(bold=True, size=13)
            ws.cell(row=row_idx, column=6).alignment = right_align
            ws.cell(row=row_idx, column=7).value = f"₹{grand_total:.2f}"
            ws.cell(row=row_idx, column=7).font = Font(bold=True, size=13)
            row_idx += 3  # Space before next order

        # Final Summary Section
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
        summary_title = ws.cell(row=row_idx, column=1)
        summary_title.value = "REPORT SUMMARY"
        summary_title.font = Font(size=14, bold=True)
        summary_title.alignment = center_align
        summary_title.fill = summary_fill
        row_idx += 2

        summary_items = [
            ("Total Orders", data['total_orders']),
            ("Total Items Sold", data['total_items']),
            ("Average Order Value", f"₹{data['avg_order']}"),
            ("TOTAL REVENUE", f"₹{data['total_revenue']}"),
        ]

        for label, value in summary_items:
            ws.cell(row=row_idx, column=5).value = label + ":"
            ws.cell(row=row_idx, column=5).font = bold_font if "TOTAL REVENUE" in label else Font(bold=True)
            ws.cell(row=row_idx, column=5).alignment = right_align
            ws.cell(row=row_idx, column=7).value = value
            ws.cell(row=row_idx, column=7).font = bold_font if "TOTAL REVENUE" in label else Font(bold=True)
            row_idx += 1

        # Column widths for readability
        column_widths = {
            'A': 6,
            'B': 28,
            'C': 18,
            'D': 8,
            'E': 18,
            'F': 20,
            'G': 18,
            'H': 14,
            'I': 14,
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Save
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        safe_period = data['period'].replace('/', '-').replace(' ', '_')
        filename = f"NoteVia_Detailed_Sales_Report_{safe_period}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
@method_decorator(login_required(login_url='signin'), name='dispatch')
@method_decorator(never_cache, name='dispatch')
class ReferralView(View):

    def get(self, request):

        if Referral.objects.filter(user=request.user).exists():
            referral = get_object_or_404(Referral, user=request.user)

        logger.info(f"refferal for : {referral.user.full_name}")


        context = {
            "user_id": request.user.id,
            "user_profile": profile(request),
            "referral": referral
        }
        return render(request, 'products/referral.html', context)
    

class WishListView(View):

    def get(self, request):
        
        wishlist = request.user.wishlist
        wishlist_items = WishlistItem.objects.filter(wishlist=wishlist, product__is_listed=True, product__category__is_listed=True)


        context = {
            "user_id": request.user.id,
            "wishlist_items": wishlist_items
        }

        return render(request, 'products/wishlist.html', context)
    
@method_decorator(login_required(login_url='signin'), name='dispatch')
@method_decorator(never_cache, name='dispatch')
class AddToWishList(View):

    def get(self, request):

        if not request.GET.get('product_id'):

            return redirect('wishlist_view')
        if request.GET.get('cart_item'):
            cart_item_id = request.GET.get('cart_item')
            cart = get_object_or_404(Cart, user=request.user)
            if CartItem.objects.filter(cart=cart, id=cart_item_id).exists():
                cart_item = get_object_or_404(CartItem, cart=cart, id=cart_item_id)
                cart_item.delete()


        product_id = request.GET.get('product_id')
        next_url = request.GET.get('next', 'shop_products')
        
        wishlist = request.user.wishlist
        
        product = Product.objects.get(id=product_id)
        if WishlistItem.objects.filter(wishlist=wishlist, product=product).exists():
            info_notify(request, "this item already exists in the wishlist")
            return redirect(next_url)
        WishlistItem.objects.get_or_create(
            wishlist=wishlist,
            product=product,
            )
        # variant = Variant.objects.get(id=variant_id)

        info_notify(request, f"the product {product.name} added to wishlist")
        return redirect(next_url)
    
@method_decorator(login_required(login_url='signin'), name='dispatch')
@method_decorator(never_cache, name='dispatch')
class RemoveToWishlist(View):
    
    def get(self, request):

        if not request.GET.get('product_id'):
            return redirect('shop_products')
        
        product_id = request.GET.get('product_id')
        next_url = request.GET.get('next', 'shop_products')
        
        wishlist = request.user.wishlist
        product = Product.objects.get(id=product_id)
        WishlistItem.objects.filter(wishlist=wishlist, product=product).delete()

        info_notify(request, f"the product {product.name} removed from wishlist")
        return redirect(next_url)
    
class WishListToCartView(View):

    def get(self, request):
        product_id = request.GET.get('product_id')
        if not product_id:
            info_notify(request, "try again..")
            return redirect('wishlist_view')
        product = get_object_or_404(Product, id=product_id )
        main_variant = product.variants.filter(is_listed=True).order_by('-stock').first()
        
        
        cart, created = Cart.objects.get_or_create(user=request.user)
        
        quantity = request.GET.get('quantity')
        if main_variant:
            if CartItem.objects.filter(cart=cart, product=product, variant=main_variant).exists():
                info_notify(request, f"this product {product.name} with same variant  is already in cart")
                return redirect('wishlist_view')
        else:
            info_notify(request, f"this product {product.name} cannot able to add cart, try again")
            return redirect('wishlist_view')
        cart_items = CartItem.objects.filter(cart=cart, is_active=True)
        cart_count = cart_items.count()
        logger.info(f"the item count is :{cart_count}")
        if cart_count >= 5:
            info_notify(request, "Please note: In cart item limit is reached")
            return redirect('wishlist_view')
        
        if not created:
            if main_variant:
                cart_item = CartItem.objects.create(cart=cart, product=product, variant=main_variant, quantity=quantity)
        else:
            if main_variant:
                cart_item = CartItem.objects.create(cart=cart, product=product, variant=main_variant, quantity=quantity)

        if cart_item:
            wishlist = request.user.wishlist
            WishlistItem.objects.filter(wishlist=wishlist, product=product).delete()
            success_notify(request, f"product {product.name} {quantity} quantity is successfully add to cart")

        return redirect('wishlist_view')

class ProductReviewView(View):

    def get(self, request):
        variant_id = request.GET.get('variant_id')
        order_id = request.GET.get('order_id')

        variant = get_object_or_404(Variant, id=variant_id)
        product  = variant.product
        review = ''
        if Review.objects.filter(user=request.user, product=product).exists():
            info_notify(request, f"already shared one review for same product {product.name} ")
            review = get_object_or_404(Review, user=request.user, product=product)
            

        context = {
            "variant_id": variant_id,
            "order_id": order_id,
            "review": review,

        }
        
        return render(request, 'products/rating_product.html', context)
    
    def post(self, request):

        variant_id = request.POST.get('variant_id')
        order_id = request.POST.get('order_id')

        if not request.POST.get('content'):
            info_notify(request, "leave your comment in the content field")
            url = f'{reverse("product_rating")}?variant_id={variant_id}&order_id={order_id}'
            return redirect(url)


        comment = request.POST.get('content').strip()
        rating = request.POST.get('rating', 1)
        

        order = get_object_or_404(Order, id=order_id)
        variant = get_object_or_404(Variant, id=variant_id)

        if Review.objects.filter(user=request.user, product=variant.product).exists():
            review = get_object_or_404(Review, user=request.user, product=variant.product)
            review.rating = rating
            review.comment = comment
            review.save()

            # reviews = Review.objects.update(user=request.user, product=variant.product, variant=variant, rating=rating, comment=comment)
        else:
            reviews = Review.objects.update_or_create(user=request.user, product=variant.product, variant=variant, rating=rating, comment=comment)
        # reviews = Review.objects.update_or_create(user=request.user, product=variant.product, variant=variant, defaults={"rating": rating, "comment": comment},)



        success_notify(request, "review added successfully")
        return redirect('order_details', order_id=order_id)
    
